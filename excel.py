#con xlsxwriter no se puede leer un archivo, sólo escribir
#con openpyxl permite leer y escribir en Excel 2010
from distutils.log import error
from plistlib import InvalidFileException
from typing import Dict

def nombreSubCarpetasListadoDeDocumentos(diccionarioArchivos,formato, acronimo,nro):

    import logging
    logger = logging.getLogger('excel.nombreSubCarpetasListadoDeDocumentos')  
    try:
        nombre = [diccionarioArchivos[formato.upper()][item]["nombre"] for item in diccionarioArchivos[formato.upper()].keys() if diccionarioArchivos[formato.upper()][item]["acronimo"].upper() == acronimo.upper()][0] 
    except:
        logger.error("Hubo un error al formar el nombre de la subcarpeta desde el Listado de documentos")
        nombre = "Carpeta"+nro
    return nombre


def dicEsqueletoDeListadoDeDocumentos(dicDocumentos = None):
    import re
    import json
    from tkinter import messagebox
    import logging
    logger = logging.getLogger('excel.dicEsqueletoDeListadoDeDocumentos')
    
    try:
        with open('json\\tipoArchivo.json',"r", encoding="utf-8") as j:
            diccionarioArchivos = json.load(j)
    except FileNotFoundError  as error:
        logger.error("Hubo un error al leer el archivo JSON: %s"%error)
        
    diccionario = dict()
    nroSubCarpeta = []
    if dicDocumentos != None:
        try:
            diccionario["carpeta02"] = {
                "nombre": "Documentación generada",
                "tipo" : "Carpeta",
                "contenido": {
                    "subcarpeta00" : {
                        "nombre":"Documentación lista para enviar",
                        "tipo": "Carpeta",
                        "contenido" : {

                        }
                    }                    
                }
            }
            
            nroArchivo = 0
            for key in dicDocumentos.keys():
                archivo = dicDocumentos[key]
                titulo = archivo["titulo"]
                formato = archivo["formato"] if archivo["formato"] else ""
                codigoTree = archivo["codigoTree"] if archivo["codigoTree"] else ""
                #armado de cantidad de subcarpetas
                if codigoTree:
                    #split para dividir por guión
                    x = re.split("-",codigoTree)
                    #en la posición 4, tomando los dos últimos dígitos es el número de la subcarpeta donde está el archivo 
                    nro = x[4][2:]
                    acronimo = x[5]
                    nombreSubCarpeta = nombreSubCarpetasListadoDeDocumentos(diccionarioArchivos,formato, acronimo, nro)
                    if nro not in nroSubCarpeta:
                        nroArchivo = 0 
                        nroSubCarpeta.append(nro)
            
                        #arma esqueleto de subcarpeta
                        diccionario["carpeta02"]["contenido"]["subcarpeta"+nro] = {"nombre" : nombreSubCarpeta, "tipo":"Carpeta","contenido":{}}
                        #arma esqueleto de primer archivo 
                        diccionario["carpeta02"]["contenido"]["subcarpeta"+nro]["contenido"]["archivo"+f"{nroArchivo:02d}"] = {"nombre": titulo, "tipo": "Archivo", "formato" : formato, "acronimo": acronimo, "codigoTree":codigoTree}
                    else:
                        #arma esqueleto de segundo o más archivo
                        diccionario["carpeta02"]["contenido"]["subcarpeta"+nro]["contenido"]["archivo"+f"{nroArchivo:02d}"] = {"nombre": titulo, "tipo": "Archivo", "formato" : formato, "acronimo": acronimo, "codigoTree":codigoTree}
                    nroArchivo += 1
                #armado de archivos dentro de subcarpeta x
            print(diccionario)
        except re.error : 
            messagebox.showerror(title = "Error", message="No se pudo obtener la estructura del Listado de Documentos. No se corresponde con el formato. El error: %s"%(re.error))
            logger.error("No se pudo obtener la estructura del Listado de Documentos. Error en split. No se corresponde con el formato. El error: %s"%(re.error))
        except IndexError as error:
            messagebox.showerror(title = "Error", message="No se pudo obtener la estructura del Listado de Documentos. Error en lectura de índice. El error: %s"%(error))
            logger.error("No se pudo obtener la estructura del Listado de Documentos. Error en lectura de índice. El error: %s"%(error))
        
        finally:    
            return diccionario if diccionario else None

#= file = 'C:\\Users\\fboz\\Desktop\\generadorDeProyectos\\prueba.xlsx' 
def lecturaExcel(fileName = None ):
    from openpyxl import load_workbook
    from tkinter import messagebox
    import logging 
    logger = logging.getLogger('excel.lecturaExcel')
    #fileName = 'C:\\Users\\fboz\\Desktop\\generadorDeProyectos\\prueba.xlsx'
    try:
        wb = load_workbook(fileName)
        hojas = wb.sheetnames
        if 'Listado de documentos' in hojas:
            hojaLD = wb['Listado de documentos']
            colNroDoc = ""
            colTitDoc = ""
            colFormato = ""
            filaInicial = -1
            for fila in hojaLD.iter_rows():
                #print("La fila es ",fila)
                for celda in fila:
                    if celda.value != None:
                        if celda.value == u'NÚMERO DE DOCUMENTO DE TREE INGENIERÍA':
                            colNroDoc = celda.coordinate[:1]
                            filaInicial = celda.row + 1
                        if celda.value == u'TÍTULO DEL DOCUMENTO':
                            colTitDoc = celda.coordinate[:1]
                            filaInicial = celda.row + 1
                        if celda.value == u'FORMATO':
                            colFormato = celda.coordinate[:1]
                            filaInicial = celda.row + 1
                        ultimaFila = celda.row
            if colNroDoc == "":
                logger.warning(u"No se encontró una columna con el nombre de'NÚMERO DE DOCUMENTO DE TREE INGENIERÍA'")
            if colTitDoc == "":
                logger.warning(u"No se encontró una columna con el nombre de'TÍTULO DEL DOCUMENTO'")
            if colFormato == "":
                logger.warning(u"No se encontró una columna con el nombre de'FORMATO'")
            if filaInicial == -1 or ultimaFila == filaInicial:
                logger.error(u"No hay filas de referencia para inicial. Faltan nombres a las columnas")
            else:
                mensaje = "El listado de documentos contiene: \n"
                dicDocumentos = dict()
                docF = 0 #f"{doc:02d}"
                for nroFila in range(filaInicial,ultimaFila+1):                        
                    docF += 1
                    if colTitDoc != "":
                        celdaTitulo = hojaLD["%s%s"%(colTitDoc,nroFila)].value
                    else:
                        celdaTitulo = None
                    if colNroDoc != "":
                        celdaNroDoc = hojaLD["%s%s"%(colNroDoc,nroFila)].value
                    else:
                        celdaNroDoc = None
                    if colFormato != "":
                        celdaFormato = hojaLD["%s%s"%(colFormato,nroFila)].value
                    else:
                        celdaFormato = None  
                    dicDocumentos.update({docF : {"titulo" : celdaTitulo, "codigoTree" : celdaNroDoc, "formato" : celdaFormato}}) #"documento%s"%(f"{docF:02d}")

                    
                    mensaje += "DOCUMENTO: %s --> TITULO: %s, Nro DOC: %s, FORMATO: %s \n"%(docF,celdaTitulo,celdaNroDoc,celdaFormato)
                logger.info(mensaje)
                #print(dicDocumentos)
        else:
            messagebox.showerror(title = "Error", message="No se encuentra la hoja: Listado de documentos. No se puede generar la carpeta deseada")
            logger.error("No se encuentra la hoja: Listado de documentos")
            
        
    except  :
        messagebox.showerror(title = "Error", message="Error abriendo el listado de documentos")
        logger.error("Error abriendo el listado de documentos")
    
    if dicDocumentos:
        print(dicDocumentos)
        diccionarioListadoDocumentos = dicEsqueletoDeListadoDeDocumentos(dicDocumentos)
    
    return diccionarioListadoDocumentos
    #hacer que si el diccionario es None, salga un aviso de que vea a crear un proyecto por defecto 
